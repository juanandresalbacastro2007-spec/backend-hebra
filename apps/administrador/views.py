# apps/administrador/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.db import connection
from django.db.models import Q
from .models import (
    Usuario, Operario, Tarea,
    AsignacionTarea, Orden, Cliente, Incidencia
)
import openpyxl
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.core.decorators import login_required_rol

# ── Decorador de protección por rol (centralizado en apps.core) ────
admin_required = login_required_rol(rol_esperado='administrador', session_key='usuario_id')


# ── Portal principal ─────────────────────────────────────────
@admin_required
def admin_portal(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])

    total_usuarios = Usuario.objects.count()
    total_clientes = Cliente.objects.count()
    total_operarios = Operario.objects.filter(estado='activo').count()
    total_ordenes = Orden.objects.count()
    ordenes_pendientes = Orden.objects.filter(estado='Pendiente').count()
    tareas_pendientes = AsignacionTarea.objects.filter(estado='Pendiente').count()

    ultimas_ordenes = Orden.objects.order_by('-fechaCreacion')[:5]
    ultimas_asignaciones = AsignacionTarea.objects.order_by('-fechaAsignacion')[:5]

    return render(request, 'administrador/admin_portal.html', {
        'usuario': usuario,
        'total_usuarios': total_usuarios,
        'total_clientes': total_clientes,
        'total_operarios': total_operarios,
        'total_ordenes': total_ordenes,
        'ordenes_pendientes': ordenes_pendientes,
        'tareas_pendientes': tareas_pendientes,
        'ultimas_ordenes': ultimas_ordenes,
        'ultimas_asignaciones': ultimas_asignaciones,
    })


# ── Usuarios ─────────────────────────────────────────────────
@admin_required
def usuarios_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    usuarios = Usuario.objects.all().order_by('rol', 'nombre')
    return render(request, 'administrador/usuarios_lista.html', {
        'usuario': usuario,
        'usuarios': usuarios,
    })


@admin_required
def usuario_crear(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        correo = request.POST.get('correoElectronico')
        contrasena = request.POST.get('contrasena')
        telefono = request.POST.get('telefono', '')
        rol = request.POST.get('rol', 'cliente')

        if Usuario.objects.filter(correoElectronico=correo).exists():
            messages.error(request, 'Ya existe un usuario con ese correo.')
            return redirect('admin_usuario_crear')

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO usuarios
                    (nombre, apellido, correoElectronico, contrasena, telefono, rol, estado)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, [nombre, apellido, correo, make_password(contrasena),
                  telefono or None, rol, 'activo'])

            id_nuevo = cursor.lastrowid

            if rol == 'cliente':
                cursor.execute("""
                    INSERT INTO clientes (idUsuario, tipoCliente, nombre, correoElectronico, estado)
                    VALUES (%s, %s, %s, %s, %s)
                """, [id_nuevo, 'Natural', f'{nombre} {apellido}', correo, 'activo'])

            elif rol == 'operario':
                especialidad = request.POST.get('especialidad', 'General')
                cursor.execute("""
                    INSERT INTO operarios (idUsuario, especialidad, fechaIngreso, estado)
                    VALUES (%s, %s, CURDATE(), %s)
                """, [id_nuevo, especialidad, 'activo'])

        messages.success(request, f'Usuario {nombre} {apellido} creado correctamente.')
        return redirect('admin_usuarios')

    return render(request, 'administrador/usuario_form.html', {
        'usuario': usuario,
        'accion': 'Crear',
    })


@admin_required
def usuario_editar(request, idUsuario):
    usuario_admin = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    usuario_editar_obj = Usuario.objects.get(idUsuario=idUsuario)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        correo = request.POST.get('correoElectronico')
        telefono = request.POST.get('telefono', '')
        rol = request.POST.get('rol')
        estado = request.POST.get('estado')

        usuario_editar_obj.nombre = nombre
        usuario_editar_obj.apellido = apellido
        usuario_editar_obj.correoElectronico = correo
        usuario_editar_obj.telefono = telefono or None
        usuario_editar_obj.rol = rol
        usuario_editar_obj.estado = estado
        usuario_editar_obj.save()

        messages.success(request, f'Usuario {nombre} actualizado correctamente.')
        return redirect('admin_usuarios')

    return render(request, 'administrador/usuario_form.html', {
        'usuario': usuario_admin,
        'usuario_editar': usuario_editar_obj,
        'accion': 'Editar',
    })


@admin_required
def usuario_eliminar(request, idUsuario):
    if request.method == 'POST':
        usuario_obj = Usuario.objects.get(idUsuario=idUsuario)
        nombre = f'{usuario_obj.nombre} {usuario_obj.apellido}'
        usuario_obj.delete()
        messages.success(request, f'Usuario {nombre} eliminado correctamente.')
    return redirect('admin_usuarios')


# ── Órdenes ──────────────────────────────────────────────────
@admin_required
def ordenes_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    ordenes = Orden.objects.all().order_by('-fechaCreacion')

    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        ordenes = ordenes.filter(
            Q(idOrden__icontains=buscar_filtro) |
            Q(idCliente__nombre__icontains=buscar_filtro) |
            Q(idCliente__empresa__icontains=buscar_filtro)
        )

    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)

    return render(request, 'administrador/ordenes_lista.html', {
        'usuario': usuario,
        'ordenes': ordenes,
        'estado_filtro': estado_filtro,
        'buscar_filtro': buscar_filtro,
    })


@admin_required
def orden_editar(request, idOrden):
    if request.method == 'POST':
        orden = get_object_or_404(Orden, pk=idOrden)
        cantidad = request.POST.get('cantidad')
        precio_unitario = request.POST.get('precio_unitario')
        fecha_entrega = request.POST.get('fecha_entrega')
        prioridad = request.POST.get('prioridad')
        estado = request.POST.get('estado')

        orden.cantidad = int(cantidad) if cantidad and cantidad.strip() else None
        orden.precioUnitario = float(precio_unitario) if precio_unitario and precio_unitario.strip() else None
        orden.fechaEntregaEstimada = fecha_entrega if fecha_entrega and fecha_entrega.strip() else None
        orden.prioridad = prioridad
        orden.estado = estado
        orden.save()
        messages.success(request, f'La orden #{idOrden} se ha modificado con éxito.')

    return redirect('admin_ordenes')


@admin_required
def orden_eliminar(request, idOrden):
    try:
        orden = get_object_or_404(Orden, pk=idOrden)
        orden.delete()
        messages.success(request, f'La orden #{idOrden} se eliminó correctamente.')
    except Exception as e:
        messages.error(request, f'Error al intentar eliminar la orden: {str(e)}')
    return redirect('admin_ordenes')


# ── Tareas ───────────────────────────────────────────────────
@admin_required
def tareas_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    asignaciones = AsignacionTarea.objects.all().order_by('-fechaAsignacion')

    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        asignaciones = asignaciones.filter(
            Q(idTarea__nombreTarea__icontains=buscar_filtro) |
            Q(idOperario__idUsuario__nombre__icontains=buscar_filtro) |
            Q(idOperario__idUsuario__apellido__icontains=buscar_filtro)
        )

    return render(request, 'administrador/tareas_lista.html', {
        'usuario': usuario,
        'asignaciones': asignaciones,
        'buscar_filtro': buscar_filtro,
    })


@admin_required
def tarea_asignar(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    operarios = Operario.objects.filter(estado='activo')
    tareas = Tarea.objects.all()

    if request.method == 'POST':
        id_tarea = request.POST.get('tarea')
        id_operario = request.POST.get('operario')
        descripcion = request.POST.get('descripcion')
        fecha_inicio = request.POST.get('fechaInicio')
        prioridad = request.POST.get('prioridad', 'Media')
        horas_estimadas = request.POST.get('horasEstimadas')

        try:
            tarea = Tarea.objects.get(idTarea=id_tarea)
            operario = Operario.objects.get(idOperario=id_operario)

            asignacion = AsignacionTarea(
                idTarea=tarea,
                idOperario=operario,
                descripcion=descripcion,
                fechaInicio=fecha_inicio if fecha_inicio and fecha_inicio.strip() else None,
                prioridad=prioridad,
                horasEstimadas=horas_estimadas if horas_estimadas and horas_estimadas.strip() else None,
                estado='Pendiente'
            )
            asignacion.save()

            messages.success(
                request,
                f'Tarea "{tarea.nombreTarea}" asignada a {operario.idUsuario.nombre} correctamente.'
            )
            return redirect('admin_tareas')

        except Exception as e:
            messages.error(request, f'Error al asignar tarea: {str(e)}')

    return render(request, 'administrador/tarea_asignar.html', {
        'usuario': usuario,
        'operarios': operarios,
        'tareas': tareas,
    })


@admin_required
def tarea_editar(request, idAsignacion):
    if request.method == 'POST':
        asignacion = get_object_or_404(AsignacionTarea, pk=idAsignacion)
        asignacion.descripcion = request.POST.get('descripcion')
        fecha_inicio = request.POST.get('fecha_inicio')
        horas_estimadas = request.POST.get('horas_estimadas')
        asignacion.fechaInicio = fecha_inicio if fecha_inicio and fecha_inicio.strip() else None
        asignacion.horasEstimadas = horas_estimadas if horas_estimadas and horas_estimadas.strip() else None
        asignacion.prioridad = request.POST.get('prioridad')
        asignacion.estado = request.POST.get('estado')
        asignacion.save()
        messages.success(request, f'La asignación #{idAsignacion} ha sido modificada con éxito.')
    return redirect('admin_tareas')


@admin_required
def tarea_eliminar(request, idAsignacion):
    try:
        asignacion = get_object_or_404(AsignacionTarea, pk=idAsignacion)
        asignacion.delete()
        messages.success(request, f'La asignación #{idAsignacion} se eliminó correctamente.')
    except Exception as e:
        messages.error(request, f'Error al intentar eliminar la asignación: {str(e)}')
    return redirect('admin_tareas')


# ── Incidencias ──────────────────────────────────────────────
@admin_required
def incidencias_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    incidencias = Incidencia.objects.all().order_by('-fechaGeneracion')

    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        incidencias = incidencias.filter(
            Q(tipoIncidencia__icontains=buscar_filtro) |
            Q(idOperario__idUsuario__nombre__icontains=buscar_filtro) |
            Q(idOperario__idUsuario__apellido__icontains=buscar_filtro)
        )

    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        incidencias = incidencias.filter(estado=estado_filtro)

    return render(request, 'administrador/incidencias_lista.html', {
        'usuario': usuario,
        'incidencias': incidencias,
        'buscar_filtro': buscar_filtro,
        'estado_filtro': estado_filtro,
    })


@admin_required
def incidencia_editar(request, idIncidencia):
    if request.method == 'POST':
        incidencia = get_object_or_404(Incidencia, pk=idIncidencia)
        incidencia.tipoIncidencia = request.POST.get('tipoIncidencia')
        incidencia.descripcion = request.POST.get('descripcion')
        incidencia.periodoEvaluado = request.POST.get('periodoEvaluado') or None
        incidencia.estado = request.POST.get('estado')
        fecha_revision = request.POST.get('fechaRevision')
        incidencia.fechaRevision = fecha_revision if fecha_revision and fecha_revision.strip() else None
        incidencia.save()
        messages.success(request, f'Incidencia #{idIncidencia} actualizada correctamente.')
    return redirect('admin_incidencias')


@admin_required
def incidencia_eliminar(request, idIncidencia):
    if request.method == 'POST':
        try:
            incidencia = get_object_or_404(Incidencia, pk=idIncidencia)
            incidencia.delete()
            messages.success(request, f'Incidencia #{idIncidencia} eliminada correctamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar la incidencia: {str(e)}')
    return redirect('admin_incidencias')


# ── Módulos / Placeholders externos ──────────────────────────
@admin_required
def produccion_placeholder(request):
    return redirect('produccion_portal')


@admin_required
def proveedores_placeholder(request):
    return redirect('proveedores')


# ── Exportar Órdenes a Excel ──────────────────────────────────
@admin_required
def exportar_ordenes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Órdenes HebraTech"

    HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    FONT_REGULAR = Font(name="Calibri", size=11)
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    headers = ["ID Orden", "Cliente / Empresa", "Fecha Creación", "Entrega Estimada",
               "Cantidad", "Precio Unitario", "Total", "Prioridad", "Estado"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ordenes = Orden.objects.all().select_related('idCliente__idUsuario').order_by('-fechaCreacion')
    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        ordenes = ordenes.filter(
            Q(idOrden__icontains=buscar_filtro) |
            Q(idCliente__nombre__icontains=buscar_filtro) |
            Q(idCliente__empresa__icontains=buscar_filtro)
        )
    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)

    for idx, orden in enumerate(ordenes, start=2):
        cliente_nombre = orden.idCliente.empresa or orden.idCliente.nombre or f"Cliente #{orden.idCliente.idCliente}"
        ws.append([
            orden.idOrden, cliente_nombre,
            orden.fechaCreacion.strftime('%Y-%m-%d') if orden.fechaCreacion else "",
            orden.fechaEntregaEstimada.strftime('%Y-%m-%d') if orden.fechaEntregaEstimada else "",
            orden.cantidad or 0,
            float(orden.precioUnitario) if orden.precioUnitario else 0,
            f"=E{idx}*F{idx}",
            orden.prioridad, orden.estado
        ])
        is_zebra = (idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = FONT_REGULAR
            cell.border = THIN_BORDER
            if is_zebra:
                cell.fill = ZEBRA_FILL
            if col_idx in [1, 3, 4, 8, 9]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif col_idx in [6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "$#,##0.00"

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="HebraTech_Reporte_Ordenes.xlsx"'
    wb.save(response)
    return response


@admin_required
def exportar_ordenes_pdf(request):
    ordenes = Orden.objects.all().select_related('idCliente__idUsuario').order_by('-fechaCreacion')
    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        ordenes = ordenes.filter(
            Q(idOrden__icontains=buscar_filtro) |
            Q(idCliente__nombre__icontains=buscar_filtro) |
            Q(idCliente__empresa__icontains=buscar_filtro)
        )
    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)

    html_string = render_to_string('administrador/ordenes_pdf.html', {'ordenes': ordenes})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="HebraTech_Reporte_Ordenes.pdf"'
    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF', status=500)
    return response